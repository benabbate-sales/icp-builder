#!/usr/bin/env python3
"""
build_territories.py

Split a graded account universe into N balanced territories within each
geographic bucket, minimising variance on both account count and total
estimated potential.

Inputs:
  --accounts   .xlsx or .csv with columns at minimum:
               FIRM_ID, FIRM_NAME, [geography column], GRADE,
               ESTIMATED_POTENTIAL, and optionally FIRM_NAME for alpha
               splits.
  --geo-col    Column name that identifies the geographic bucket to
               split within (e.g. 'SUB_REGION' or 'STATE_TERRITORY_BUCKET').
               The script builds balanced splits *inside each bucket*;
               it does not re-bucket accounts across buckets.
  --splits     Path to a .csv describing how many territories each
               bucket should be split into, and optionally the split
               method. Columns:
                   BUCKET,            N_TERRITORIES,  METHOD,      NAME_PREFIX
                   "UKI & Nordics",   2,              "alpha",     "Territory"
                   "BENELUX",         1,              "-",         "BENELUX"
                   "DACH",            1,              "-",         "DACH"
               METHOD is one of:
                   "-"       — bucket is one territory, no split
                   "balance" — greedy balance on count + potential
                   "alpha"   — split by FIRM_NAME first letter into N groups
  --out        Output .xlsx or .csv. Adds a TERRITORY column.

Notes:
  - Greedy balance uses a simple longest-processing-time heuristic on
    ESTIMATED_POTENTIAL (assign the next biggest account to the lightest
    territory). This produces tight balance for N up to ~5.
  - Alpha split picks first-letter boundaries that divide FIRM_NAME list
    as evenly as possible. For N=2 this typically lands near 'I/J'.
  - Accounts with GRADE == 'UNGRADED' are assigned TERRITORY = 'UNGRADED'
    and excluded from balancing.
  - Accounts missing the geo column value are assigned TERRITORY =
    'UNCOVERED'.

Typical run:
    python build_territories.py \
        --accounts universe_graded.xlsx \
        --geo-col  SUB_REGION \
        --splits   splits.csv \
        --out      universe_with_territories.xlsx
"""

import argparse
import csv
import os
import sys
from collections import defaultdict


def _load_table(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            return reader.fieldnames or [], rows
    if ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl  # type: ignore
        except ImportError:
            sys.exit("openpyxl required. pip install openpyxl")
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else "" for h in next(it)]
        rows = []
        for r in it:
            if all(c is None for c in r):
                continue
            rows.append({h: ("" if v is None else v) for h, v in zip(headers, r)})
        return headers, rows
    sys.exit(f"Unsupported extension: {ext}")


def _write_table(path, headers, rows):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=headers)
            w.writeheader()
            for r in rows:
                w.writerow({h: r.get(h, "") for h in headers})
        return
    if ext in (".xlsx", ".xlsm"):
        import openpyxl  # type: ignore
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        wb.save(path)
        return
    sys.exit(f"Unsupported extension: {ext}")


def _to_float(v):
    if v in (None, ""):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _balance_greedy(accounts, n, prefix):
    """Longest-processing-time split on ESTIMATED_POTENTIAL."""
    buckets = [{"name": f"{prefix} {i+1}", "accounts": [], "pot": 0.0} for i in range(n)]
    sorted_accts = sorted(accounts, key=lambda a: _to_float(a.get("ESTIMATED_POTENTIAL", 0)),
                          reverse=True)
    for a in sorted_accts:
        b = min(buckets, key=lambda x: (x["pot"], len(x["accounts"])))
        b["accounts"].append(a)
        b["pot"] += _to_float(a.get("ESTIMATED_POTENTIAL", 0))
    assignments = {}
    for b in buckets:
        for a in b["accounts"]:
            assignments[a["_idx"]] = b["name"]
    return assignments


def _alpha_split(accounts, n, prefix):
    """Split accounts by FIRM_NAME first letter into N groups of ~equal size."""
    sorted_accts = sorted(accounts, key=lambda a: str(a.get("FIRM_NAME", "")).upper())
    total = len(sorted_accts)
    targets = [round(total * (i + 1) / n) for i in range(n)]
    assignments = {}
    cur = 0
    for i, cutoff in enumerate(targets):
        group = sorted_accts[cur:cutoff]
        if not group:
            name = f"{prefix} {i+1}"
        else:
            first = str(group[0].get("FIRM_NAME", ""))[:1].upper() or "A"
            last = str(group[-1].get("FIRM_NAME", ""))[:1].upper() or "Z"
            name = f"{prefix} {i+1} ({first}-{last})"
        for a in group:
            assignments[a["_idx"]] = name
        cur = cutoff
    return assignments


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--accounts", required=True)
    ap.add_argument("--geo-col", required=True)
    ap.add_argument("--splits", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    a_headers, a_rows = _load_table(args.accounts)
    s_headers, s_rows = _load_table(args.splits)

    if args.geo_col not in a_headers:
        sys.exit(f"Column '{args.geo_col}' not found in accounts file.")

    splits = {}
    for r in s_rows:
        bucket = str(r.get("BUCKET", "")).strip()
        if not bucket:
            continue
        splits[bucket] = {
            "n": int(r.get("N_TERRITORIES", 1) or 1),
            "method": (str(r.get("METHOD", "-")).strip() or "-").lower(),
            "prefix": str(r.get("NAME_PREFIX", bucket)).strip() or bucket,
        }

    # Index rows so assignments can be stable.
    for i, r in enumerate(a_rows):
        r["_idx"] = i

    by_bucket = defaultdict(list)
    for r in a_rows:
        bucket = str(r.get(args.geo_col, "")).strip()
        grade = str(r.get("GRADE", "")).strip().upper()
        if grade == "UNGRADED":
            r["TERRITORY"] = "UNGRADED"
            continue
        if not bucket:
            r["TERRITORY"] = "UNCOVERED"
            continue
        by_bucket[bucket].append(r)

    unassigned_buckets = []
    for bucket, accts in by_bucket.items():
        cfg = splits.get(bucket)
        if not cfg:
            unassigned_buckets.append(bucket)
            for a in accts:
                a["TERRITORY"] = "UNASSIGNED"
            continue
        n = cfg["n"]
        method = cfg["method"]
        prefix = cfg["prefix"]
        if n <= 1 or method == "-":
            for a in accts:
                a["TERRITORY"] = prefix
            continue
        if method == "balance":
            m = _balance_greedy(accts, n, prefix)
        elif method == "alpha":
            m = _alpha_split(accts, n, prefix)
        else:
            sys.exit(f"Unknown split method: {method}")
        for a in accts:
            a["TERRITORY"] = m[a["_idx"]]

    # Remove helper.
    for r in a_rows:
        r.pop("_idx", None)

    out_headers = list(a_headers)
    if "TERRITORY" not in out_headers:
        out_headers.append("TERRITORY")
    _write_table(args.out, out_headers, a_rows)

    # Summary.
    summary = defaultdict(lambda: {"n": 0, "pot": 0.0})
    for r in a_rows:
        t = r.get("TERRITORY", "")
        summary[t]["n"] += 1
        summary[t]["pot"] += _to_float(r.get("ESTIMATED_POTENTIAL", 0))

    print("Territory summary:")
    for t, s in sorted(summary.items(), key=lambda kv: -kv[1]["pot"]):
        print(f"  {t:40s}  accounts={s['n']:4d}  potential=${s['pot']:,.0f}")
    if unassigned_buckets:
        print("\nBuckets with no split config (accounts marked UNASSIGNED):")
        for b in unassigned_buckets:
            print(f"  - {b}")


if __name__ == "__main__":
    main()
