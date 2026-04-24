#!/usr/bin/env python3
"""
grade_accounts.py

Apply a per-segment scoring profile to an account universe and write back
GRADE for each account.

Inputs:
  --accounts   Path to an .xlsx or .csv with the account universe. Must
               contain columns: FIRM_ID, GRADING_SEGMENT, and one column
               per scoring dimension used by that segment's profile.
  --profile    Path to a scoring-profile .csv. One row per
               (SEGMENT, dimension combination) -> GRADE mapping. The
               dimension columns in the profile must match the dimension
               columns in the accounts file (column names are matched
               case-insensitively and ignoring leading/trailing whitespace).
  --out        Path to write the enriched accounts file. Same format as
               the input (.xlsx or .csv). A GRADE column is added (or
               overwritten if it already exists).

Behaviour:
  - Any account whose (segment, dimensions) tuple has no match in the
    profile is graded 'UNGRADED' and flagged. The script prints a list
    of unmatched combinations at the end so you can fix the profile.
  - Dimension value comparisons are case-insensitive and whitespace-
    stripped, so 'Y' and 'y ' compare equal.
  - The profile can have extra columns (e.g. NOTES) - they are ignored.

Typical run:
    python grade_accounts.py \
        --accounts universe.xlsx \
        --profile  scoring_profile.csv \
        --out      universe_graded.xlsx
"""

import argparse
import csv
import os
import sys
from collections import defaultdict


def _load_table(path):
    """Return (headers, list_of_row_dicts). Supports .csv and .xlsx."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            return reader.fieldnames or [], rows
    if ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl  # type: ignore
        except ImportError:
            sys.exit("openpyxl is required for .xlsx files. pip install openpyxl")
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else "" for h in next(rows_iter)]
        rows = []
        for r in rows_iter:
            if all(c is None for c in r):
                continue
            rows.append({h: ("" if v is None else v) for h, v in zip(headers, r)})
        return headers, rows
    sys.exit(f"Unsupported file extension: {ext}")


def _write_table(path, headers, rows):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for r in rows:
                writer.writerow({h: r.get(h, "") for h in headers})
        return
    if ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl  # type: ignore
        except ImportError:
            sys.exit("openpyxl is required for .xlsx files. pip install openpyxl")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        wb.save(path)
        return
    sys.exit(f"Unsupported file extension: {ext}")


def _norm(v):
    return str(v).strip().upper() if v is not None else ""


def _build_lookup(profile_rows, profile_headers):
    """segment_name -> { (dim_col, dim_col, ...) : [dim_values_tuple -> grade] }."""
    # Identify which profile columns are dimension columns. Everything except
    # SEGMENT, GRADE, NOTES, and blank headers is treated as a dimension.
    reserved = {"SEGMENT", "GRADE", "NOTES", ""}
    dim_cols = [h for h in profile_headers if _norm(h) not in {_norm(x) for x in reserved}]

    by_segment = defaultdict(dict)  # segment_upper -> {dim_tuple -> grade}
    segment_dims = {}               # segment_upper -> list of dim col names (canonical)
    for row in profile_rows:
        seg = _norm(row.get("SEGMENT", ""))
        if not seg or seg.startswith("#"):
            continue
        grade = _norm(row.get("GRADE", ""))
        if grade not in {"A", "B", "C"}:
            continue
        dim_tuple = tuple(_norm(row.get(c, "")) for c in dim_cols)
        by_segment[seg][dim_tuple] = grade
        segment_dims.setdefault(seg, dim_cols)
    return by_segment, segment_dims


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--accounts", required=True)
    ap.add_argument("--profile", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    a_headers, a_rows = _load_table(args.accounts)
    p_headers, p_rows = _load_table(args.profile)

    by_segment, segment_dims = _build_lookup(p_rows, p_headers)
    if not by_segment:
        sys.exit("Scoring profile appears empty after parsing. Check --profile file.")

    # Ensure GRADE column exists in the output headers.
    out_headers = list(a_headers)
    if "GRADE" not in out_headers:
        out_headers.append("GRADE")

    unmatched = defaultdict(int)
    graded_counts = defaultdict(int)
    for row in a_rows:
        seg = _norm(row.get("GRADING_SEGMENT", ""))
        dims = segment_dims.get(seg)
        if not dims:
            row["GRADE"] = "UNGRADED"
            unmatched[(seg, "<no profile for segment>")] += 1
            continue
        key = tuple(_norm(row.get(c, "")) for c in dims)
        grade = by_segment[seg].get(key)
        if grade is None:
            row["GRADE"] = "UNGRADED"
            unmatched[(seg, key)] += 1
        else:
            row["GRADE"] = grade
            graded_counts[grade] += 1

    _write_table(args.out, out_headers, a_rows)

    total = sum(graded_counts.values())
    print(f"Graded {total} accounts: "
          f"A={graded_counts['A']}, B={graded_counts['B']}, C={graded_counts['C']}")
    if unmatched:
        print(f"\n{sum(unmatched.values())} accounts could not be graded. "
              "Unmatched combinations:")
        for (seg, combo), n in sorted(unmatched.items(), key=lambda kv: -kv[1]):
            if combo == "<no profile for segment>":
                print(f"  [{seg}] {n} accounts — segment has no profile")
            else:
                print(f"  [{seg}] {combo} — {n} accounts")
        print("\nAdd these combinations to the scoring profile and rerun.")


if __name__ == "__main__":
    main()
